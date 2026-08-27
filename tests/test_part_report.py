"""Regression tests for the fixed-column part report workflow.

Covers the things that would be expensive to discover in production: the
value-resolution policy (user input is never overwritten), full multi-page
traversal, part attribution without cross-contamination, and the Excel header
contract.

The vision model is stubbed, so these run offline with no API key and no cost.

Run with either:
    python tests/test_part_report.py
    pytest tests/test_part_report.py
"""

from __future__ import annotations

import io
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

# --------------------------------------------------------------------------
# Stub the vision model before anything imports it.
# --------------------------------------------------------------------------

from app.pipeline import gemini_client  # noqa: E402

PAGE_RESPONSES = {
    "MOUNTING BRACKET": {
        "page_summary": "Detail drawing of mounting bracket",
        "page_type": "detail",
        "parts_on_page": [
            {
                "part_no": "BR-1042",
                "matched_user_part_no": "BR-1042",
                "match_confidence": 0.97,
                "fields": {
                    "description": {"value": "Mounting Bracket", "confidence": 0.95,
                                    "source_text": "DESCRIPTION MOUNTING BRACKET"},
                    # Deliberately disagrees with the user's "6".
                    "thickness": {"value": "8", "unit": "mm", "confidence": 0.91,
                                  "source_text": "8 THK"},
                    "weight_kg": {"value": "2.40", "unit": "kg", "confidence": 0.93,
                                  "source_text": "WEIGHT 2.40 kg"},
                    "process": {"value": "Laser Cut + Bending", "confidence": 0.9,
                                "source_text": "PROCESS LASER CUT + BENDING"},
                    "length": {"value": "200", "unit": "mm", "confidence": 0.88,
                               "source_text": "OVERALL SIZE 200 x 100 x 45"},
                    "width": {"value": "100", "unit": "mm", "confidence": 0.88,
                              "source_text": "OVERALL SIZE 200 x 100 x 45"},
                    # Below the confidence threshold on purpose.
                    "height": {"value": "45", "unit": "mm", "confidence": 0.30,
                               "source_text": "OVERALL SIZE 200 x 100 x 45"},
                },
            }
        ],
        "bom_parts": [
            {"s_no": "1", "part_no": "BR-1042", "description": "MOUNTING BRACKET",
             "dwg_no": "BA120V220102", "quantity": "1", "confidence": 0.95},
            {"s_no": "2", "part_no": "CP-2071", "description": "COVER PLATE",
             "dwg_no": "BA120V220101", "quantity": "2", "confidence": 0.95},
        ],
        "findings": [
            {"category": "hole", "value": "4 x M8 THRU", "part_no": "BR-1042", "confidence": 0.9},
            {"category": "gdt", "value": "PERPENDICULARITY 0.2 | A", "part_no": "BR-1042", "confidence": 0.86},
            {"category": "weld", "value": "FILLET WELD 5mm ALL ROUND", "part_no": "BR-1042", "confidence": 0.8},
            {"category": "material", "value": "MS IS2062 GR.B", "part_no": "BR-1042", "confidence": 0.94},
            {"category": "other", "value": "STRAY NOTE", "part_no": "ZZ-9999", "confidence": 0.6},
        ],
    },
    "COVER PLATE": {
        "page_summary": "Cover plate detail",
        "page_type": "detail",
        "parts_on_page": [
            {
                # Punctuation differs from the supplied "CP-2071" on purpose.
                "part_no": "CP2071",
                "matched_user_part_no": None,
                "match_confidence": 0.8,
                "fields": {
                    "description": {"value": "Cover Plate", "confidence": 0.94, "source_text": "COVER PLATE"},
                    "thickness": {"value": "3", "unit": "mm", "confidence": 0.92, "source_text": "3 THK"},
                    "weight_kg": {"value": "0.85", "unit": "kg", "confidence": 0.9, "source_text": "0.85 kg"},
                },
            }
        ],
        "findings": [
            {"category": "hole", "value": "6 x DIA 6.5 THRU", "part_no": "CP-2071", "confidence": 0.9},
            {"category": "gdt", "value": "FLATNESS 0.1", "part_no": "CP-2071", "confidence": 0.88},
        ],
    },
    "SOLE DETAIL PART": {
        "page_summary": "Single detail part, no parts list",
        "page_type": "detail",
        "bom_parts": [],
        "title_block": {"part_no": "SD-9001", "description": "SPACER RING",
                        "dwg_no": "D-9001", "confidence": 0.9},
        "parts_on_page": [{
            "part_no": "SD-9001", "matched_user_part_no": None, "match_confidence": 0.9,
            "fields": {"thickness": {"value": "6", "unit": "mm", "confidence": 0.9,
                                     "source_text": "6 THK"}},
        }],
        "findings": [],
    },
    "GENERAL NOTES": {
        "page_summary": "General notes",
        "page_type": "notes",
        "parts_on_page": [
            {
                "part_no": "BR-1042",
                "matched_user_part_no": "BR-1042",
                "match_confidence": 0.9,
                "fields": {"thickness": {"value": "8", "unit": "mm", "confidence": 0.7,
                                         "source_text": "REV C - THICKNESS CHANGED 5 TO 6"}},
            }
        ],
        "findings": [
            {"category": "note", "value": "ALL DIMENSIONS IN MILLIMETRES", "confidence": 0.95},
        ],
    },
}


class StubGemini:
    """Answers per page by matching a marker in the prompt's OCR block."""

    sdk = "stub"
    available = True

    def __init__(self, *args, **kwargs):
        self.calls: list[str] = []

    def is_available(self) -> bool:
        return self.available

    @property
    def unavailable_reason(self) -> str:
        return "" if self.available else "GEMINI_API_KEY is not set"

    def generate_json(self, prompt, images=None, **kwargs):
        self.calls.append(prompt)
        for marker, response in PAGE_RESPONSES.items():
            if marker in prompt:
                return response
        return {"parts_on_page": [], "findings": []}


STUB = StubGemini()
gemini_client.GeminiClient = lambda *a, **k: STUB

from app.backend.excel_report import ExcelReportWriter  # noqa: E402
from app.engine.report_resolver import ReportResolver  # noqa: E402
from app.models.part_schemas import (  # noqa: E402
    REPORT_COLUMNS,
    FieldEvidence,
    PartInput,
    ValueStatus,
)
from app.pipeline.part_pipeline import PartReportPipeline  # noqa: E402

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"


# --------------------------------------------------------------------------
# Fixture: a synthetic 3-page drawing
# --------------------------------------------------------------------------


def build_test_pdf(path: Path) -> Path:
    """Write a 3-page drawing with title blocks, notes and geometry."""
    import fitz

    pages = [
        ("MOUNTING BRACKET", "BR-1042",
         [("PART NO", "BR-1042"), ("DESCRIPTION", "MOUNTING BRACKET"),
          ("MATERIAL", "MS IS2062 GR.B"), ("THICKNESS", "8 THK"),
          ("WEIGHT", "2.40 kg"), ("PROCESS", "LASER CUT + BENDING")],
         ["OVERALL SIZE 200 x 100 x 45", "4 x M8 THRU TAPPED HOLES",
          "GD&T: PERPENDICULARITY 0.2 | A", "FILLET WELD 5mm ALL ROUND"]),
        ("COVER PLATE", "CP-2071",
         [("PART NO", "CP-2071"), ("DESCRIPTION", "COVER PLATE"),
          ("MATERIAL", "SS 304"), ("THICKNESS", "3 THK"), ("WEIGHT", "0.85 kg")],
         ["OVERALL SIZE 150 x 90 x 3", "6 x DIA 6.5 THRU HOLES", "FLATNESS 0.1"]),
        ("GENERAL NOTES & REVISION", "",
         [("SHEET", "3 OF 3"), ("DRAWN BY", "R. KUMAR")],
         ["1. ALL DIMENSIONS IN MILLIMETRES",
          "REV C - THICKNESS CHANGED 5 TO 6 (BR-1042)"]),
    ]

    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    for title, part_no, block, notes in pages:
        page = doc.new_page(width=842, height=595)
        page.draw_rect(fitz.Rect(15, 15, 827, 580), color=(0, 0, 0), width=1.5)
        page.insert_text((30, 45), title, fontsize=16, fontname="hebo")
        if part_no:
            page.insert_text((30, 68), "PART No: " + part_no, fontsize=12, fontname="hebo")
        page.draw_rect(fitz.Rect(60, 100, 360, 280), color=(0, 0, 0), width=1)
        y = 320
        for note in notes:
            page.insert_text((60, y), note, fontsize=9)
            y += 17
        ty = 400
        for label, value in block:
            page.insert_text((528, ty), label, fontsize=8, fontname="hebo")
            page.insert_text((648, ty), value, fontsize=8)
            ty += 22
    doc.save(str(path))
    doc.close()
    return path


def get_pdf() -> Path:
    path = FIXTURE_DIR / "synthetic_drawing.pdf"
    if not path.exists():
        build_test_pdf(path)
    return path


_result_cache: dict = {}


def analysed():
    """Run the pipeline once and reuse the result across tests."""
    if "result" not in _result_cache:
        STUB.available = True
        parts = [
            PartInput(s_no="1", part_no="BR-1042", thickness="6"),
            PartInput(s_no="2", part_no="CP-2071"),
        ]
        _result_cache["result"] = PartReportPipeline().run(
            get_pdf(), parts, document_id="testdoc"
        )
    return _result_cache["result"]


# --------------------------------------------------------------------------
# Resolution policy
# --------------------------------------------------------------------------


def test_user_value_is_never_overwritten():
    resolver = ReportResolver(confidence_threshold=0.4)
    part = PartInput(part_no="X1", thickness="6")
    evidence = {"X1": [FieldEvidence(field="thickness", value="8", unit="mm",
                                     page_number=2, confidence=0.95, source_text="8 THK")]}
    cell = resolver.resolve([part], evidence, [])[0][0].cells["THICKNESS"]

    assert cell.value == "6", "the user's entry must survive a high-confidence disagreement"
    assert cell.status == ValueStatus.CONFLICT
    assert cell.drawing_value == "8 mm", "the drawing reading must still be recorded"
    assert cell.page_references == [2]


def test_blank_is_filled_from_drawing():
    resolver = ReportResolver(confidence_threshold=0.4)
    part = PartInput(part_no="X1")
    evidence = {"X1": [FieldEvidence(field="description", value="Bracket",
                                     page_number=1, confidence=0.9, source_text="D")]}
    cell = resolver.resolve([part], evidence, [])[0][0].cells["DESCRIPTION"]

    assert cell.value == "Bracket"
    assert cell.status == ValueStatus.FILLED_FROM_DRAWING


def test_low_confidence_is_not_detected_rather_than_guessed():
    resolver = ReportResolver(confidence_threshold=0.4)
    part = PartInput(part_no="X1")
    evidence = {"X1": [FieldEvidence(field="length", value="200",
                                     page_number=1, confidence=0.2, source_text="L")]}
    cell = resolver.resolve([part], evidence, [])[0][0].cells["LENGTH (mm)"]

    assert cell.value == "Not Detected", "a low-confidence reading must not reach the report"
    assert cell.status == ValueStatus.MISSING
    assert cell.drawing_value == "200", "but it must still be visible for audit"


def test_missing_everywhere_is_not_detected():
    resolver = ReportResolver()
    cell = resolver.resolve([PartInput(part_no="X1")], {}, [])[0][0].cells["WIDTH (mm)"]
    assert cell.value == "Not Detected"


def test_equivalent_values_do_not_raise_a_false_conflict():
    resolver = ReportResolver()
    assert resolver._values_agree("6", "6.0 mm")
    assert resolver._values_agree("Laser Cutting", "Laser Cut")
    assert not resolver._values_agree("6", "8 mm")
    assert not resolver._values_agree("Laser Cut", "Machining")


# --------------------------------------------------------------------------
# Pipeline
# --------------------------------------------------------------------------


def test_every_page_is_analysed():
    result = analysed()
    assert result.total_pages == 3
    assert result.pages_analyzed == 3, "the pipeline must not stop at the first page"
    assert len(result.page_analyses) == 3


def test_evidence_is_merged_across_pages():
    row = next(r for r in analysed().rows if r.part_no == "BR-1042")
    assert row.cells["THICKNESS"].page_references == [1, 3], (
        "thickness appears on pages 1 and 3 and both must be cited"
    )


def test_part_numbers_match_despite_punctuation():
    row = next(r for r in analysed().rows if r.part_no == "CP-2071")
    # Data reached the row at all, which is the point: the drawing prints
    # "CP2071" while the operator supplied "CP-2071".
    assert row.cells["DESCRIPTION"].value.upper() == "COVER PLATE", (
        "'CP2071' on the drawing must match supplied 'CP-2071'"
    )
    assert row.cells["DWG NO"].value == "BA120V220101", (
        "the per-part drawing number must come from the BOM row"
    )


def test_findings_are_not_mixed_between_parts():
    result = analysed()
    bracket = {f.value for f in next(r for r in result.rows if r.part_no == "BR-1042").findings}
    plate = {f.value for f in next(r for r in result.rows if r.part_no == "CP-2071").findings}

    assert "4 x M8 THRU" in bracket
    assert "FLATNESS 0.1" in plate
    assert not bracket & plate, "no finding may be attributed to both parts"


def test_unattributable_findings_are_quarantined():
    values = {f.value for f in analysed().unmatched_findings}
    assert "STRAY NOTE" in values, "data for an unsupplied part must not join a row"
    assert "ALL DIMENSIONS IN MILLIMETRES" in values


def test_pipeline_degrades_without_a_vision_model():
    STUB.available = False
    try:
        result = PartReportPipeline().run(
            get_pdf(), [PartInput(s_no="1", part_no="BR-1042", thickness="6")]
        )
        assert len(result.rows) == 1
        assert result.rows[0].cells["THICKNESS"].value == "6", "typed values still survive"
        assert result.rows[0].cells["LENGTH (mm)"].value == "Not Detected"
        assert any("GEMINI_API_KEY" in w for w in result.warnings), (
            "the user must be told why nothing was extracted"
        )
    finally:
        STUB.available = True


# --------------------------------------------------------------------------
# Discovery mode: empty grid -> rows from the drawing's BOM
# --------------------------------------------------------------------------


def test_empty_grid_builds_rows_from_the_bom_table():
    """Clicking Analyze with nothing typed must still produce a full report."""
    STUB.available = True
    result = PartReportPipeline().run(get_pdf(), [], document_id="discovery")

    assert result.discovery_mode is True
    assert len(result.rows) == 2, "both BOM rows must become report rows"

    by_part = {row.part_no: row for row in result.rows}
    assert set(by_part) == {"BR-1042", "CP-2071"}

    bracket = by_part["BR-1042"]
    assert bracket.discovered is True
    assert bracket.cells["DESCRIPTION"].value.upper() == "MOUNTING BRACKET"
    assert bracket.cells["DWG NO"].value == "BA120V220102"
    assert bracket.cells["S No"].value == "1"
    # Per-part technical data still lands on the discovered row.
    assert bracket.cells["THICKNESS"].value == "8 mm"
    assert any("parts-list table" in w for w in result.warnings)

    # Page 2 prints "CP2071" while the BOM row key is "CP-2071". Matching on
    # the raw string would silently drop every field found for that part.
    plate = by_part["CP-2071"]
    assert plate.cells["THICKNESS"].value == "3 mm", (
        "evidence must reach a row whose part number is punctuated differently"
    )
    assert plate.cells["WEIGHT (IN KG)"].value == "0.85"


def test_discovered_rows_are_not_treated_as_operator_input():
    """A BOM value must never masquerade as something a human confirmed."""
    from app.models.part_schemas import ValueSource

    STUB.available = True
    result = PartReportPipeline().run(get_pdf(), [], document_id="discovery2")
    cell = result.rows[0].cells["DESCRIPTION"]
    assert cell.source == ValueSource.DRAWING
    assert cell.user_value is None


def test_supplied_rows_win_over_the_bom_table():
    """Typing one part must not cause the other 16 to be added behind you."""
    STUB.available = True
    result = PartReportPipeline().run(
        get_pdf(), [PartInput(s_no="1", part_no="BR-1042", thickness="6")],
        document_id="supplied",
    )
    assert result.discovery_mode is False
    assert len(result.rows) == 1, "the BOM must not add rows the operator did not ask for"
    assert result.rows[0].cells["THICKNESS"].value == "6"


# --------------------------------------------------------------------------
# Generality: no parts-list table at all
# --------------------------------------------------------------------------


def _single_part_pdf() -> Path:
    """A one-page drawing with no BOM - only a title block."""
    import fitz

    path = FIXTURE_DIR / "single_detail.pdf"
    if path.exists():
        return path
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    page = doc.new_page(width=842, height=595)
    page.insert_text((40, 45), "SOLE DETAIL PART", fontsize=16, fontname="hebo")
    page.insert_text((60, 120), "PART NO SD-9001", fontsize=10)
    page.insert_text((60, 140), "DESCRIPTION SPACER RING", fontsize=10)
    page.insert_text((60, 160), "6 THK", fontsize=10)
    doc.save(str(path))
    doc.close()
    return path


def test_drawing_with_no_parts_list_still_produces_a_row():
    """Not every drawing has a BOM. A single detail sheet must still work."""
    STUB.available = True
    result = PartReportPipeline().run(_single_part_pdf(), [], document_id="single")

    assert result.discovery_mode is True
    assert len(result.rows) == 1, "the title block must supply the row"

    cells = result.rows[0].cells
    assert cells["PART NO"].value == "SD-9001"
    assert cells["DESCRIPTION"].value == "SPACER RING"
    assert cells["DWG NO"].value == "D-9001"
    assert cells["THICKNESS"].value == "6 mm"
    assert any("title block" in w for w in result.warnings)


def test_process_vocabulary_is_not_restricted():
    """The report must carry whatever process the drawing states."""
    from app.pipeline import part_extractor as pe

    assert "not a list to choose from" in pe.FIELD_GUIDANCE, (
        "the prompt must not present its examples as a closed vocabulary"
    )


def test_process_markers_for_bending_and_welding():
    """Bending is recognised from a bend radius/angle and welding from weld
    symbols/fillet-size callouts, so the process gets captured."""
    from app.pipeline import part_extractor as pe

    guidance = " ".join(pe.FIELD_GUIDANCE.split()).upper()
    page_prompt = " ".join(pe.PAGE_PROMPT.split()).upper()

    assert "BEND RADIUS" in guidance or "BENDING" in guidance
    assert "BEND ANGLE" in guidance or "90°" in guidance
    assert ("FILET" in guidance) or ("FILLET" in guidance) or ("Z1" in guidance) or (
        "WELD" in guidance
    )
    assert "WELDING" in guidance
    assert "BENDING" in page_prompt and "WELDING" in page_prompt


def test_process_auto_detected_from_drawing_markers():
    """The PROCESS field is auto-generated from bending/welding markers."""
    from app.pipeline.process_detector import detect_processes_from_text

    # Bending: bend radius/angle callouts.
    assert "bending" in detect_processes_from_text("BEND R5 90 DEG")
    assert "bending" in detect_processes_from_text("BEND ALLOWANCE 2")
    # Welding: fillet/weld symbols such as z1, z2, a5.
    assert "welding" in detect_processes_from_text("WELD NOTE, FILLET Z1 Z2")
    assert "welding" in detect_processes_from_text("FILLET WELD A5")
    # Known processes from their words.
    assert "laser cutting" in detect_processes_from_text("laser cutting")
    assert "tapping" in detect_processes_from_text("TAP M4")
    # No guesswork from ordinary dimensions / labels.
    assert detect_processes_from_text("R12 dimension only") == []
    assert detect_processes_from_text("4 holes M8 counterbore") == []


def test_orientation_rules_are_in_the_prompt():
    """Development-view ban, width-opposite-length, height-from-Z-axis."""
    from app.pipeline import part_extractor as pe

    flat = " ".join(pe.FIELD_GUIDANCE.split()).upper()
    assert "DEVELOPMENT" in flat and "FLAT PATTERN" in flat, (
        "the prompt must forbid development/flat-pattern dimensions"
    )
    assert "OPPOSITE" in flat and "PERPENDICULAR" in flat, (
        "width must be perpendicular/opposite to length"
    )
    page = " ".join(pe.PAGE_PROMPT.split()).upper()
    assert "Z AXIS" in page, "height must be tied to the 3D model's Z axis"


def test_development_view_dimensions_never_fill_lwh():
    """A development/flat-pattern reading is excluded from LENGTH/WIDTH/HEIGHT."""
    from app.pipeline import part_extractor as pe

    extractor = pe.PartExtractor()
    parsed = {
        "bom_parts": [],
        "title_block": None,
        "parts_on_page": [
            {
                "part_no": "BR-1042",
                "matched_user_part_no": "BR-1042",
                "fields": {
                    "length": {"value": "1240", "unit": "mm", "confidence": 0.9,
                               "source_text": "DEVELOPMENT LENGTH 1240"},
                    "width": {"value": "200", "unit": "mm", "confidence": 0.85,
                              "source_text": "OVERALL 200"},
                    "height": {"value": "45", "unit": "mm", "confidence": 0.9,
                               "source_text": "ISOMETRIC VIEW Z 45"},
                },
            }
        ],
        "findings": [],
        "part_numbers": [],
    }
    parts = [PartInput(s_no="1", part_no="BR-1042")]
    out = extractor._parse_page_response(parsed, 1, parts)

    fields = {f.field for _, f in out["evidence"]}
    assert "length" not in fields, "development-view length must not reach the report"
    assert fields == {"width", "height"}
    assert any("Development" in (f.detail or "") for f in out["findings"]), (
        "the excluded reading must stay visible as a finding"
    )


def test_bom_headings_are_matched_on_meaning_not_exact_text():
    import re

    from app.pipeline import part_extractor as pe

    flat = re.sub(r"\s+", " ", pe.PAGE_PROMPT)
    for token in ("BILL OF MATERIALS", "ITEM LIST", "POS", "PART CODE"):
        assert token in flat, f"prompt should anticipate '{token}' headings"
    assert "Match on meaning, not on exact column headings" in flat


# --------------------------------------------------------------------------
# Excel contract
# --------------------------------------------------------------------------


def workbook():
    if "wb" not in _result_cache:
        path = ExcelReportWriter().generate(
            analysed(), output_path=FIXTURE_DIR / "out" / "report.xlsx"
        )
        _result_cache["wb"] = openpyxl.load_workbook(path)
    return _result_cache["wb"]


def test_headers_are_fixed_and_ordered():
    ws = workbook()["Report"]
    headers = [ws.cell(row=2, column=i + 1).value for i in range(len(REPORT_COLUMNS))]
    assert headers == list(REPORT_COLUMNS)


def test_report_sheet_is_formatted():
    ws = workbook()["Report"]
    assert ws.freeze_panes == "A3", "the header row must stay visible while scrolling"
    assert ws.auto_filter.ref is not None
    assert "ENGINEERING DRAWING ANALYSIS REPORT" in str(ws["A1"].value)
    assert ws.column_dimensions["C"].width == 40
    assert ws.cell(row=3, column=1).border.left.style == "thin"
    assert ws.page_setup.orientation == "landscape"


def test_numeric_columns_hold_real_numbers():
    ws = workbook()["Report"]
    assert ws.cell(row=3, column=6).value == 6, "thickness must be a number, not text"
    assert ws.cell(row=3, column=5).value == 2.4
    assert ws.cell(row=3, column=5).number_format == "0.00"
    length = ws.cell(row=3, column=8)
    assert length.value == 200, "LENGTH (mm) must be a number, not text"


def test_unresolved_values_read_not_detected():
    ws = workbook()["Report"]
    assert ws.cell(row=3, column=10).value == "Not Detected"


def test_conflicts_are_visible_in_the_workbook():
    ws = workbook()["Report"]
    cell = ws.cell(row=3, column=6)
    assert cell.fill.start_color.rgb.endswith("FFF2CC"), "conflict cells are shaded amber"
    assert cell.comment is not None, "and carry the drawing value on hover"


def test_supporting_sheets_exist():
    wb = workbook()
    assert wb.sheetnames == ["Report", "Traceability", "Drawing Information", "Analysis Log"]
    statuses = [wb["Traceability"].cell(row=r, column=5).value
                for r in range(2, wb["Traceability"].max_row + 1)]
    assert "conflict" in statuses
    assert "filled" in statuses
    assert wb["Analysis Log"].cell(row=4, column=1).value == 3, "all 3 pages logged"


# --------------------------------------------------------------------------
# API
# --------------------------------------------------------------------------


def test_api_round_trip():
    from fastapi.testclient import TestClient

    import app.backend.part_routes as part_routes

    part_routes._pipeline = PartReportPipeline()
    from app.backend.api import app as fastapi_app

    client = TestClient(fastapi_app)

    assert client.get("/").status_code == 200, "the UI is served from the API root"
    caps = client.get("/api/part-report/capabilities").json()
    assert caps["report_columns"] == list(REPORT_COLUMNS)

    with open(get_pdf(), "rb") as handle:
        upload = client.post(
            "/api/part-report/upload",
            files=[("files", ("drawing.pdf", handle, "application/pdf"))],
        )
    assert upload.status_code == 200, upload.text
    document_id = upload.json()["document_id"]
    assert upload.json()["page_count"] == 3
    assert upload.json()["file_count"] == 1

    # Guards that stop ambiguous input reaching the pipeline.
    assert client.post("/api/part-report/analyze", json={
        "document_id": document_id, "parts": [{"description": "no part number"}]
    }).status_code == 400
    assert client.post("/api/part-report/analyze", json={
        "document_id": document_id, "parts": [{"part_no": "A-1"}, {"part_no": "a1"}]
    }).status_code == 400, "duplicate part numbers would make attribution ambiguous"

    started = client.post("/api/part-report/analyze", json={
        "document_id": document_id,
        "parts": [{"s_no": "1", "part_no": "BR-1042", "thickness": "6"},
                  {"s_no": "2", "part_no": "CP-2071"}],
    })
    assert started.status_code == 200, started.text
    job_id = started.json()["job_id"]

    # Generous: with PaddleOCR installed a page takes seconds, and abandoning
    # the poll would leave a worker thread running into the next test.
    progress = {}
    for _ in range(600):
        progress = client.get(f"/api/part-report/progress/{job_id}").json()
        if progress["status"] in {"complete", "error"}:
            break
        time.sleep(0.5)
    assert progress["status"] == "complete", progress

    payload = client.get(f"/api/part-report/result/{job_id}").json()
    assert payload["table"]["columns"] == list(REPORT_COLUMNS)
    assert len(payload["table"]["rows"]) == 2
    assert payload["stats"]["conflicts"] == 1

    excel = client.get(f"/api/part-report/excel/{job_id}")
    assert excel.status_code == 200
    assert excel.content[:2] == b"PK", "must be a real xlsx archive"
    downloaded = openpyxl.load_workbook(io.BytesIO(excel.content))
    assert [downloaded["Report"].cell(row=2, column=i + 1).value
            for i in range(len(REPORT_COLUMNS))] == list(REPORT_COLUMNS)

    image = client.get(f"/api/part-report/page-image/{document_id}/2")
    assert image.status_code == 200 and image.content[:4] == b"\x89PNG"


# --------------------------------------------------------------------------
# Legacy endpoint emits the same fixed headers
# --------------------------------------------------------------------------


def test_legacy_analysis_projects_onto_the_same_fixed_columns():
    """The analysis-only endpoint must not produce a second Excel format."""
    from app.backend.legacy_adapter import build_report_from_analysis
    from app.models.schemas import (
        BOMItem,
        DocumentAnalysisResult,
        ExtractionCategory,
        PageResult,
    )

    page = PageResult(page_number=1, page_width=2480, page_height=1754)
    page.bom_items.append(
        BOMItem(value="1", category=ExtractionCategory.BOM, page_number=1,
                confidence=0.9, part_number="BR-1042",
                description="MOUNTING BRACKET", weight="2400 g")
    )
    analysis = DocumentAnalysisResult(filename="d.pdf", file_path="d.pdf", total_pages=1)
    analysis.page_results = [page]

    report = build_report_from_analysis(analysis)
    assert list(report.rows[0].cells.keys()) == list(REPORT_COLUMNS)

    cells = report.rows[0].cells
    assert cells["PART NO"].value == "BR-1042"
    assert cells["DESCRIPTION"].value == "MOUNTING BRACKET"
    assert cells["WEIGHT (IN KG)"].value == "2.4 kg", "grams must convert to kg"
    assert cells["DWG NO"].value == "Not Detected"
    # Fields the legacy model cannot carry stay honest rather than invented.
    assert cells["THICKNESS"].value == "Not Detected"
    assert cells["PROCESS"].value == "Not Detected"


def test_legacy_weight_without_a_unit_is_not_assumed_to_be_kg():
    from app.backend.legacy_adapter import _normalise_weight

    assert _normalise_weight("2400 g") == "2.4 kg"
    assert _normalise_weight("2.4 kg") == "2.4 kg"
    assert _normalise_weight("5 lb") == "2.26796 kg"
    assert _normalise_weight("2400") == "2400", "no printed unit means no conversion"
    assert _normalise_weight(None) is None


# --------------------------------------------------------------------------
# Runner (so the suite works without pytest installed)
# --------------------------------------------------------------------------


def main() -> int:
    tests = [(name, obj) for name, obj in sorted(globals().items())
             if name.startswith("test_") and callable(obj)]
    failures = []

    for name, test in tests:
        try:
            test()
            print(f"  PASS  {name}")
        except Exception as exc:  # noqa: BLE001 - this is the test runner
            print(f"  FAIL  {name}: {exc}")
            failures.append(name)

    print("-" * 62)
    if failures:
        print(f"{len(failures)} of {len(tests)} failed: {', '.join(failures)}")
        return 1
    print(f"All {len(tests)} tests passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
