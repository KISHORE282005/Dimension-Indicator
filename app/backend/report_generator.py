"""Report generation module — PDF and Excel output.

Generates professional engineering analysis reports from DocumentAnalysisResult.
Primary output: Excel with Dimension Control List (DCL) format.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.config import settings
from app.models.schemas import (
    DimensionControlRow,
    DocumentAnalysisResult,
    ExtractionCategory,
    SeverityLevel,
)

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates PDF and Excel reports from analysis results."""

    def __init__(self) -> None:
        self.output_dir = settings.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ==================================================================
    # EXCEL REPORT — Primary download format
    # ==================================================================

    def generate_excel(
        self,
        result: DocumentAnalysisResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Generate Excel report with Dimension Control List as primary sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, falling back to JSON")
            return self.generate_json(result, output_path)

        if output_path is None:
            output_path = self.output_dir / f"{result.document_id}_report.xlsx"

        wb = openpyxl.Workbook()

        # --- Style constants ---
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        critical_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        non_critical_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        critical_font = Font(color="991B1B", bold=True, size=10, name="Calibri")
        non_critical_font = Font(color="065F46", bold=True, size=10, name="Calibri")
        data_font = Font(size=10, name="Calibri")
        bold_font = Font(bold=True, size=10, name="Calibri")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def style_header(ws, row: int = 1, col_count: int = 5) -> None:
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

        def style_data_cell(ws, row: int, col: int, alignment=None) -> None:
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border
            cell.alignment = alignment or left_align

        # ==============================================================
        # SHEET 1: Dimension Control List (Primary)
        # ==============================================================
        ws_dcl = wb.active
        ws_dcl.title = "Dimension Control List"
        ws_dcl.sheet_properties.tabColor = "1A1A2E"

        # Title row
        ws_dcl.merge_cells("A1:F1")
        title_cell = ws_dcl["A1"]
        title_cell.value = "ENGINEERING DRAWING — DIMENSION CONTROL LIST"
        title_cell.font = Font(bold=True, size=14, color="1A1A2E", name="Calibri")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_dcl.merge_cells("A2:F2")
        ws_dcl["A2"].value = f"Drawing: {result.filename} | Pages: {result.total_pages} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws_dcl["A2"].font = Font(size=9, color="666666", name="Calibri")
        ws_dcl["A2"].alignment = Alignment(horizontal="center")

        # Headers at row 4
        dcl_headers = [
            "Dimension No.",
            "Specification",
            "Criticality",
            "Mode of Control",
            "Mode of Inspection",
            "Page",
        ]
        for col, header in enumerate(dcl_headers, 1):
            ws_dcl.cell(row=4, column=col, value=header)
        style_header(ws_dcl, row=4, col_count=len(dcl_headers))

        # Build consolidated list
        result.build_consolidated_dimension_control()
        dcl_rows = result.consolidated_dimension_control

        if not dcl_rows:
            # Fallback: build from raw items
            dcl_rows = self._build_dcl_from_raw(result)

        # Data rows
        row_num = 5
        for dcl_row in dcl_rows:
            ws_dcl.cell(row=row_num, column=1, value=dcl_row.dimension_number)
            ws_dcl.cell(row=row_num, column=2, value=dcl_row.specification)
            ws_dcl.cell(row=row_num, column=3, value=dcl_row.criticality)
            ws_dcl.cell(row=row_num, column=4, value=dcl_row.mode_of_control)
            ws_dcl.cell(row=row_num, column=5, value=dcl_row.mode_of_inspection)
            ws_dcl.cell(row=row_num, column=6, value=dcl_row.page_number or "-")

            for col in range(1, 7):
                style_data_cell(ws_dcl, row_num, col, center_align if col != 2 else left_align)

            # Color-code criticality
            crit_cell = ws_dcl.cell(row=row_num, column=3)
            if dcl_row.criticality == "Critical":
                crit_cell.fill = critical_fill
                crit_cell.font = critical_font
            else:
                crit_cell.fill = non_critical_fill
                crit_cell.font = non_critical_font

            row_num += 1

        # Column widths
        ws_dcl.column_dimensions["A"].width = 14
        ws_dcl.column_dimensions["B"].width = 40
        ws_dcl.column_dimensions["C"].width = 16
        ws_dcl.column_dimensions["D"].width = 18
        ws_dcl.column_dimensions["E"].width = 20
        ws_dcl.column_dimensions["F"].width = 8

        # Summary row
        row_num += 1
        ws_dcl.cell(row=row_num, column=1, value="TOTAL").font = bold_font
        ws_dcl.cell(row=row_num, column=2, value=f"{len(dcl_rows)} dimensions/features").font = bold_font

        critical_count = sum(1 for r in dcl_rows if r.criticality == "Critical")
        non_critical_count = len(dcl_rows) - critical_count
        ws_dcl.cell(row=row_num, column=3, value=f"{critical_count}C / {non_critical_count}NC").font = bold_font
        for col in range(1, 7):
            ws_dcl.cell(row=row_num, column=col).border = border

        # Freeze panes
        ws_dcl.freeze_panes = "A5"

        # ==============================================================
        # SHEET 2: Dimension Details
        # ==============================================================
        all_dims = [d for pr in result.page_results for d in pr.dimensions]
        if all_dims:
            ws_dim = wb.create_sheet("Dimension Details")
            ws_dim.sheet_properties.tabColor = "0F3460"
            dim_headers = [
                "Dim No.", "Value", "Type", "Nominal", "Unit",
                "Upper Limit", "Lower Limit", "Tolerance",
                "Specification", "Criticality", "Control", "Inspection",
                "Page", "Confidence", "Source",
            ]
            for col, header in enumerate(dim_headers, 1):
                ws_dim.cell(row=1, column=col, value=header)
            style_header(ws_dim, row=1, col_count=len(dim_headers))

            for idx, d in enumerate(all_dims, 1):
                r = idx + 1
                ws_dim.cell(row=r, column=1, value=idx)
                ws_dim.cell(row=r, column=2, value=str(d.value))
                ws_dim.cell(row=r, column=3, value=d.dimension_type or "-")
                ws_dim.cell(row=r, column=4, value=d.nominal_value)
                ws_dim.cell(row=r, column=5, value=d.unit or "-")
                ws_dim.cell(row=r, column=6, value=d.upper_limit)
                ws_dim.cell(row=r, column=7, value=d.lower_limit)
                ws_dim.cell(row=r, column=8, value=d.tolerance_value)
                ws_dim.cell(row=r, column=9, value=d.specification or "-")
                ws_dim.cell(row=r, column=10, value=d.criticality.value)
                ws_dim.cell(row=r, column=11, value=d.mode_of_control.value)
                ws_dim.cell(row=r, column=12, value=d.mode_of_inspection.value)
                ws_dim.cell(row=r, column=13, value=d.page_number)
                ws_dim.cell(row=r, column=14, value=f"{d.confidence:.0%}")
                ws_dim.cell(row=r, column=15, value=d.source_type.value if hasattr(d.source_type, 'value') else str(d.source_type))
                for col in range(1, 16):
                    style_data_cell(ws_dim, r, col)
            ws_dim.freeze_panes = "A2"

        # ==============================================================
        # SHEET 3: Hole Details
        # ==============================================================
        all_holes = [h for pr in result.page_results for h in pr.holes]
        if all_holes:
            ws_hole = wb.create_sheet("Hole Details")
            ws_hole.sheet_properties.tabColor = "E94560"
            hole_headers = [
                "No.", "Hole Type", "Diameter", "Depth", "Thread Spec",
                "Quantity", "Specification", "Criticality",
                "Control", "Inspection", "Page", "Confidence",
            ]
            for col, header in enumerate(hole_headers, 1):
                ws_hole.cell(row=1, column=col, value=header)
            style_header(ws_hole, row=1, col_count=len(hole_headers))

            for idx, h in enumerate(all_holes, 1):
                r = idx + 1
                ws_hole.cell(row=r, column=1, value=idx)
                ws_hole.cell(row=r, column=2, value=h.hole_type or "-")
                ws_hole.cell(row=r, column=3, value=h.diameter)
                ws_hole.cell(row=r, column=4, value=h.depth)
                ws_hole.cell(row=r, column=5, value=h.thread_spec or "-")
                ws_hole.cell(row=r, column=6, value=h.quantity)
                ws_hole.cell(row=r, column=7, value=h.specification or "-")
                ws_hole.cell(row=r, column=8, value=h.criticality.value)
                ws_hole.cell(row=r, column=9, value=h.mode_of_control.value)
                ws_hole.cell(row=r, column=10, value=h.mode_of_inspection.value)
                ws_hole.cell(row=r, column=11, value=h.page_number)
                ws_hole.cell(row=r, column=12, value=f"{h.confidence:.0%}")
                for col in range(1, 13):
                    style_data_cell(ws_hole, r, col)
            ws_hole.freeze_panes = "A2"

        # ==============================================================
        # SHEET 4: Tolerance Details
        # ==============================================================
        all_tols = [t for pr in result.page_results for t in pr.tolerances]
        if all_tols:
            ws_tol = wb.create_sheet("Tolerance Details")
            ws_tol.sheet_properties.tabColor = "F59E0B"
            tol_headers = [
                "No.", "Nominal", "Upper Tol", "Lower Tol",
                "Upper Limit", "Lower Limit", "Fit Class",
                "Specification", "Criticality", "Control", "Inspection",
                "Page", "Confidence",
            ]
            for col, header in enumerate(tol_headers, 1):
                ws_tol.cell(row=1, column=col, value=header)
            style_header(ws_tol, row=1, col_count=len(tol_headers))

            for idx, t in enumerate(all_tols, 1):
                r = idx + 1
                ws_tol.cell(row=r, column=1, value=idx)
                ws_tol.cell(row=r, column=2, value=t.nominal_value)
                ws_tol.cell(row=r, column=3, value=t.upper_tolerance)
                ws_tol.cell(row=r, column=4, value=t.lower_tolerance)
                ws_tol.cell(row=r, column=5, value=t.upper_limit)
                ws_tol.cell(row=r, column=6, value=t.lower_limit)
                ws_tol.cell(row=r, column=7, value=t.fit_class or "-")
                ws_tol.cell(row=r, column=8, value=t.specification or "-")
                ws_tol.cell(row=r, column=9, value=t.criticality.value)
                ws_tol.cell(row=r, column=10, value=t.mode_of_control.value)
                ws_tol.cell(row=r, column=11, value=t.mode_of_inspection.value)
                ws_tol.cell(row=r, column=12, value=t.page_number)
                ws_tol.cell(row=r, column=13, value=f"{t.confidence:.0%}")
                for col in range(1, 14):
                    style_data_cell(ws_tol, r, col)
            ws_tol.freeze_panes = "A2"

        # ==============================================================
        # SHEET 5: GD&T
        # ==============================================================
        all_gdt = [g for pr in result.page_results for g in pr.gd_t_items]
        if all_gdt:
            ws_gdt = wb.create_sheet("GD&T")
            ws_gdt.sheet_properties.tabColor = "3B82F6"
            gdt_headers = [
                "No.", "Characteristic", "Symbol", "Tolerance Value",
                "Zone Shape", "Modifier", "Primary Datum",
                "Secondary Datum", "Tertiary Datum", "Feature Control Frame",
                "Page", "Confidence",
            ]
            for col, header in enumerate(gdt_headers, 1):
                ws_gdt.cell(row=1, column=col, value=header)
            style_header(ws_gdt, row=1, col_count=len(gdt_headers))

            for idx, g in enumerate(all_gdt, 1):
                r = idx + 1
                ws_gdt.cell(row=r, column=1, value=idx)
                ws_gdt.cell(row=r, column=2, value=g.characteristic or "-")
                ws_gdt.cell(row=r, column=3, value=g.symbol or "-")
                ws_gdt.cell(row=r, column=4, value=g.tolerance_value)
                ws_gdt.cell(row=r, column=5, value=g.tolerance_zone_shape or "-")
                ws_gdt.cell(row=r, column=6, value=g.modifier or "-")
                ws_gdt.cell(row=r, column=7, value=g.primary_datum or "-")
                ws_gdt.cell(row=r, column=8, value=g.secondary_datum or "-")
                ws_gdt.cell(row=r, column=9, value=g.tertiary_datum or "-")
                ws_gdt.cell(row=r, column=10, value=g.feature_control_frame or "-")
                ws_gdt.cell(row=r, column=11, value=g.page_number)
                ws_gdt.cell(row=r, column=12, value=f"{g.confidence:.0%}")
                for col in range(1, 13):
                    style_data_cell(ws_gdt, r, col)
            ws_gdt.freeze_panes = "A2"

        # ==============================================================
        # SHEET 6: Welding
        # ==============================================================
        all_welds = [w for pr in result.page_results for w in pr.welding_items]
        if all_welds:
            ws_weld = wb.create_sheet("Welding")
            ws_weld.sheet_properties.tabColor = "EF4444"
            weld_headers = [
                "No.", "Weld Type", "Size", "Length", "Joint Type",
                "Arrow Side", "Other Side", "Contour", "Field Weld",
                "Page", "Confidence",
            ]
            for col, header in enumerate(weld_headers, 1):
                ws_weld.cell(row=1, column=col, value=header)
            style_header(ws_weld, row=1, col_count=len(weld_headers))

            for idx, w in enumerate(all_welds, 1):
                r = idx + 1
                ws_weld.cell(row=r, column=1, value=idx)
                ws_weld.cell(row=r, column=2, value=w.weld_type or "-")
                ws_weld.cell(row=r, column=3, value=w.weld_size or "-")
                ws_weld.cell(row=r, column=4, value=w.weld_length or "-")
                ws_weld.cell(row=r, column=5, value=w.joint_type or "-")
                ws_weld.cell(row=r, column=6, value="Yes" if w.arrow_side else "No")
                ws_weld.cell(row=r, column=7, value="Yes" if w.other_side else "No")
                ws_weld.cell(row=r, column=8, value=w.contour or "-")
                ws_weld.cell(row=r, column=9, value="Yes" if w.field_weld else "No")
                ws_weld.cell(row=r, column=10, value=w.page_number)
                ws_weld.cell(row=r, column=11, value=f"{w.confidence:.0%}")
                for col in range(1, 12):
                    style_data_cell(ws_weld, r, col)
            ws_weld.freeze_panes = "A2"

        # ==============================================================
        # SHEET 7: BOM / Parts
        # ==============================================================
        all_bom = [b for pr in result.page_results for b in pr.bom_items]
        if all_bom:
            ws_bom = wb.create_sheet("BOM")
            ws_bom.sheet_properties.tabColor = "8B5CF6"
            bom_headers = [
                "Item", "Part Number", "Description", "Quantity",
                "Material", "Weight", "Remarks", "Page",
            ]
            for col, header in enumerate(bom_headers, 1):
                ws_bom.cell(row=1, column=col, value=header)
            style_header(ws_bom, row=1, col_count=len(bom_headers))

            for idx, b in enumerate(all_bom, 1):
                r = idx + 1
                ws_bom.cell(row=r, column=1, value=b.row_index or idx)
                ws_bom.cell(row=r, column=2, value=b.part_number or "-")
                ws_bom.cell(row=r, column=3, value=b.description or "-")
                ws_bom.cell(row=r, column=4, value=b.quantity)
                ws_bom.cell(row=r, column=5, value=b.material or "-")
                ws_bom.cell(row=r, column=6, value=b.weight or "-")
                ws_bom.cell(row=r, column=7, value=b.remarks or "-")
                ws_bom.cell(row=r, column=8, value=b.page_number)
                for col in range(1, 9):
                    style_data_cell(ws_bom, r, col)
            ws_bom.freeze_panes = "A2"

        # ==============================================================
        # SHEET 8: Surface Finish
        # ==============================================================
        all_sf = [s for pr in result.page_results for s in pr.surface_finishes]
        if all_sf:
            ws_sf = wb.create_sheet("Surface Finish")
            ws_sf.sheet_properties.tabColor = "10B981"
            sf_headers = [
                "No.", "Roughness Value", "Unit", "Method",
                "Lay Symbol", "Original Spec", "Page", "Confidence",
            ]
            for col, header in enumerate(sf_headers, 1):
                ws_sf.cell(row=1, column=col, value=header)
            style_header(ws_sf, row=1, col_count=len(sf_headers))

            for idx, s in enumerate(all_sf, 1):
                r = idx + 1
                ws_sf.cell(row=r, column=1, value=idx)
                ws_sf.cell(row=r, column=2, value=s.roughness_value)
                ws_sf.cell(row=r, column=3, value=s.roughness_unit or "Ra")
                ws_sf.cell(row=r, column=4, value=s.surface_method or "-")
                ws_sf.cell(row=r, column=5, value=s.lay_symbol or "-")
                ws_sf.cell(row=r, column=6, value=s.original_specification or str(s.value))
                ws_sf.cell(row=r, column=7, value=s.page_number)
                ws_sf.cell(row=r, column=8, value=f"{s.confidence:.0%}")
                for col in range(1, 9):
                    style_data_cell(ws_sf, r, col)
            ws_sf.freeze_panes = "A2"

        # ==============================================================
        # SHEET 9: Materials
        # ==============================================================
        all_mat = [m for pr in result.page_results for m in pr.materials]
        if all_mat:
            ws_mat = wb.create_sheet("Materials")
            ws_mat.sheet_properties.tabColor = "78716C"
            mat_headers = [
                "No.", "Material Spec", "Name", "Grade",
                "Condition", "Standard", "Page",
            ]
            for col, header in enumerate(mat_headers, 1):
                ws_mat.cell(row=1, column=col, value=header)
            style_header(ws_mat, row=1, col_count=len(mat_headers))

            for idx, m in enumerate(all_mat, 1):
                r = idx + 1
                ws_mat.cell(row=r, column=1, value=idx)
                ws_mat.cell(row=r, column=2, value=m.material_spec or "-")
                ws_mat.cell(row=r, column=3, value=m.material_name or "-")
                ws_mat.cell(row=r, column=4, value=m.material_grade or "-")
                ws_mat.cell(row=r, column=5, value=m.condition or "-")
                ws_mat.cell(row=r, column=6, value=m.standard or "-")
                ws_mat.cell(row=r, column=7, value=m.page_number)
                for col in range(1, 8):
                    style_data_cell(ws_mat, r, col)
            ws_mat.freeze_panes = "A2"

        # ==============================================================
        # SHEET 10: Manufacturing Notes
        # ==============================================================
        all_notes = [n for pr in result.page_results for n in pr.manufacturing_notes]
        if all_notes:
            ws_notes = wb.create_sheet("Notes")
            ws_notes.sheet_properties.tabColor = "F97316"
            note_headers = ["Note No.", "Type", "Text", "Page"]
            for col, header in enumerate(note_headers, 1):
                ws_notes.cell(row=1, column=col, value=header)
            style_header(ws_notes, row=1, col_count=len(note_headers))

            for idx, n in enumerate(all_notes, 1):
                r = idx + 1
                ws_notes.cell(row=r, column=1, value=n.note_number or idx)
                ws_notes.cell(row=r, column=2, value=n.note_type or "-")
                ws_notes.cell(row=r, column=3, value=n.note_text or str(n.value))
                ws_notes.cell(row=r, column=4, value=n.page_number)
                for col in range(1, 5):
                    style_data_cell(ws_notes, r, col)
            ws_notes.freeze_panes = "A2"

        # ==============================================================
        # SHEET 11: Issues
        # ==============================================================
        if result.all_issues:
            ws_issues = wb.create_sheet("Issues")
            ws_issues.sheet_properties.tabColor = "DC2626"
            issue_headers = [
                "#", "Severity", "Issue Type", "Description",
                "Page", "Recommendation",
            ]
            for col, header in enumerate(issue_headers, 1):
                ws_issues.cell(row=1, column=col, value=header)
            style_header(ws_issues, row=1, col_count=len(issue_headers))

            for idx, issue in enumerate(result.all_issues, 1):
                r = idx + 1
                ws_issues.cell(row=r, column=1, value=idx)
                ws_issues.cell(row=r, column=2, value=issue.severity.value.upper())
                ws_issues.cell(row=r, column=3, value=issue.issue_type.value)
                ws_issues.cell(row=r, column=4, value=issue.description)
                ws_issues.cell(row=r, column=5, value=issue.page_number or "-")
                ws_issues.cell(row=r, column=6, value=issue.recommendation or "-")
                for col in range(1, 7):
                    style_data_cell(ws_issues, r, col)
                # Color-code severity
                sev_cell = ws_issues.cell(row=r, column=2)
                if issue.severity in (SeverityLevel.ERROR, SeverityLevel.CRITICAL):
                    sev_cell.font = Font(color="991B1B", bold=True, size=10)
                elif issue.severity == SeverityLevel.WARNING:
                    sev_cell.font = Font(color="92400E", bold=True, size=10)
            ws_issues.freeze_panes = "A2"

        # ==============================================================
        # SHEET 12: Summary
        # ==============================================================
        ws_sum = wb.create_sheet("Summary")
        ws_sum.sheet_properties.tabColor = "6366F1"
        result.build_summary()

        ws_sum.cell(row=1, column=1, value="Engineering Drawing Analysis Summary").font = Font(
            bold=True, size=14, color="1A1A2E"
        )
        ws_sum.cell(row=3, column=1, value="Filename").font = bold_font
        ws_sum.cell(row=3, column=2, value=result.filename)
        ws_sum.cell(row=4, column=1, value="Total Pages").font = bold_font
        ws_sum.cell(row=4, column=2, value=result.total_pages)
        ws_sum.cell(row=5, column=1, value="Processing Time").font = bold_font
        ws_sum.cell(row=5, column=2, value=f"{result.total_processing_time_seconds:.1f}s")

        valid_str = "PASS" if (result.validation_result and result.validation_result.is_valid) else "FAIL"
        ws_sum.cell(row=6, column=1, value="Validation Status").font = bold_font
        ws_sum.cell(row=6, column=2, value=valid_str)

        ws_sum.cell(row=7, column=1, value="Total Issues").font = bold_font
        ws_sum.cell(row=7, column=2, value=len(result.all_issues))

        ws_sum.cell(row=9, column=1, value="Category").font = bold_font
        ws_sum.cell(row=9, column=2, value="Count").font = bold_font
        for idx, (k, v) in enumerate(result.extraction_summary.items(), 10):
            ws_sum.cell(row=idx, column=1, value=k.replace("_", " ").title())
            ws_sum.cell(row=idx, column=2, value=v)

        ws_sum.column_dimensions["A"].width = 25
        ws_sum.column_dimensions["B"].width = 40

        # Save
        wb.save(str(output_path))
        logger.info("Excel report generated: %s", output_path)
        return output_path

    # ==================================================================
    # PDF REPORT
    # ==================================================================

    def generate_pdf(
        self,
        result: DocumentAnalysisResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Generate a professional PDF engineering analysis report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
                PageBreak,
                HRFlowable,
            )
        except ImportError:
            logger.warning("reportlab not installed, generating text report instead")
            return self._generate_text_report(result, output_path)

        if output_path is None:
            output_path = self.output_dir / f"{result.document_id}_report.pdf"

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=20 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            "SectionTitle",
            parent=styles["Heading1"],
            fontSize=13,
            spaceAfter=6,
            textColor=colors.HexColor("#1a1a2e"),
        ))
        styles.add(ParagraphStyle(
            "SubSection",
            parent=styles["Heading2"],
            fontSize=10,
            spaceAfter=4,
            textColor=colors.HexColor("#16213e"),
        ))
        styles.add(ParagraphStyle(
            "NormalText",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
        ))

        elements = []

        # Title
        elements.append(Paragraph(
            "Engineering Drawing Analysis — Dimension Control Report",
            styles["Title"],
        ))
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(
            f"<b>File:</b> {result.filename} | "
            f"<b>Pages:</b> {result.total_pages} | "
            f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            styles["NormalText"],
        ))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))

        # Executive Summary
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph("1. Executive Summary", styles["SectionTitle"]))
        result.build_summary()
        valid_str = "PASS" if (result.validation_result and result.validation_result.is_valid) else "FAIL"
        summary_text = (
            f"<b>{result.filename}</b> — {result.total_pages} pages analyzed in "
            f"{result.total_processing_time_seconds:.1f}s. "
            f"Validation: <b>{valid_str}</b> — "
            f"{len(result.all_issues)} issues found."
        )
        elements.append(Paragraph(summary_text, styles["NormalText"]))

        # ====== DIMENSION CONTROL LIST (Primary table) ======
        elements.append(PageBreak())
        elements.append(Paragraph("2. Dimension Control List", styles["SectionTitle"]))

        result.build_consolidated_dimension_control()
        dcl_rows = result.consolidated_dimension_control
        if not dcl_rows:
            dcl_rows = self._build_dcl_from_raw(result)

        if dcl_rows:
            dcl_table_data = [["Dim No.", "Specification", "Criticality", "Mode of Control", "Mode of Inspection", "Page"]]
            for row in dcl_rows:
                dcl_table_data.append([
                    str(row.dimension_number),
                    row.specification[:50],
                    row.criticality,
                    row.mode_of_control,
                    row.mode_of_inspection,
                    str(row.page_number or "-"),
                ])

            col_widths = [50, 220, 70, 90, 100, 40]
            t = Table(dcl_table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 0), (2, -1), "CENTER"),
                ("ALIGN", (5, 0), (5, -1), "CENTER"),
            ]
            # Color critical rows
            for i, row in enumerate(dcl_rows, 1):
                if row.criticality == "Critical":
                    style_cmds.append(("BACKGROUND", (2, i), (2, i), colors.HexColor("#FEE2E2")))
            t.setStyle(TableStyle(style_cmds))
            elements.append(t)

            # Summary
            elements.append(Spacer(1, 3 * mm))
            critical_count = sum(1 for r in dcl_rows if r.criticality == "Critical")
            elements.append(Paragraph(
                f"<b>Total:</b> {len(dcl_rows)} items | "
                f"<b>Critical:</b> {critical_count} | "
                f"<b>Non-Critical:</b> {len(dcl_rows) - critical_count}",
                styles["NormalText"],
            ))

        # Drawing Info
        for pr in result.page_results:
            if pr.drawing_metadata:
                elements.append(Spacer(1, 4 * mm))
                elements.append(Paragraph(
                    f"3. Drawing Information — Page {pr.page_number}",
                    styles["SectionTitle"],
                ))
                dm = pr.drawing_metadata
                info_rows = []
                for label, val in [
                    ("Drawing Number", dm.drawing_number),
                    ("Revision", dm.revision),
                    ("Title", dm.title),
                    ("Scale", dm.scale),
                    ("Drawn By", dm.drawn_by),
                    ("Checked By", dm.checked_by),
                    ("Date", dm.date),
                    ("Company", dm.company),
                ]:
                    if val:
                        info_rows.append([label, str(val)])
                if info_rows:
                    t = Table([["Field", "Value"]] + info_rows, colWidths=[100, 250])
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(t)

        # Manufacturing Notes
        all_notes = [n for pr in result.page_results for n in pr.manufacturing_notes]
        if all_notes:
            elements.append(PageBreak())
            elements.append(Paragraph("4. Manufacturing Notes", styles["SectionTitle"]))
            for n in all_notes:
                elements.append(Paragraph(
                    f"<b>Note {n.note_number or ''}:</b> {n.note_text} <i>(Page {n.page_number})</i>",
                    styles["NormalText"],
                ))
                elements.append(Spacer(1, 1 * mm))

        # Issues
        if result.all_issues:
            elements.append(Spacer(1, 4 * mm))
            elements.append(Paragraph("5. Issues and Warnings", styles["SectionTitle"]))
            issue_data = [["#", "Severity", "Type", "Description", "Page"]]
            for idx, issue in enumerate(result.all_issues, 1):
                issue_data.append([
                    str(idx),
                    issue.severity.value.upper(),
                    issue.issue_type.value,
                    issue.description[:60],
                    str(issue.page_number or "-"),
                ])
            t = Table(issue_data, colWidths=[25, 50, 90, 200, 35])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B0000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(t)

        # Build PDF
        doc.build(elements)
        logger.info("PDF report generated: %s", output_path)
        return output_path

    # ==================================================================
    # JSON EXPORT
    # ==================================================================

    def generate_json(
        self,
        result: DocumentAnalysisResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Export the full analysis result as JSON."""
        if output_path is None:
            output_path = self.output_dir / f"{result.document_id}_result.json"
        output_path.write_text(result.model_dump_json(indent=2), encoding="utf-8")
        logger.info("JSON export: %s", output_path)
        return output_path

    # ==================================================================
    # FALLBACK: Text report
    # ==================================================================

    def _generate_text_report(
        self,
        result: DocumentAnalysisResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        if output_path is None:
            output_path = self.output_dir / f"{result.document_id}_report.txt"
        lines = [
            "=" * 70,
            "ENGINEERING DRAWING — DIMENSION CONTROL REPORT",
            "=" * 70,
            f"File: {result.filename}",
            f"Pages: {result.total_pages}",
            f"Time: {result.total_processing_time_seconds:.1f}s",
            f"Valid: {result.validation_result.is_valid if result.validation_result else 'N/A'}",
            "",
            "-" * 70,
            "DIMENSION CONTROL LIST",
            "-" * 70,
            f"{'Dim No.':<10} {'Specification':<40} {'Criticality':<15} {'Control':<15} {'Inspection':<15}",
            "-" * 70,
        ]
        result.build_consolidated_dimension_control()
        for row in result.consolidated_dimension_control:
            lines.append(
                f"{row.dimension_number:<10} {row.specification[:38]:<40} {row.criticality:<15} {row.mode_of_control:<15} {row.mode_of_inspection:<15}"
            )
        lines.append("-" * 70)
        lines.append(f"Total: {len(result.consolidated_dimension_control)} items")
        lines.append("")
        lines.append("ISSUES:")
        for issue in result.all_issues:
            lines.append(f"  [{issue.severity.value}] {issue.issue_type.value}: {issue.description}")
        output_path.write_text("\n".join(lines), encoding="utf-8")
        return output_path

    # ==================================================================
    # HELPERS
    # ==================================================================

    @staticmethod
    def _build_dcl_from_raw(result: DocumentAnalysisResult) -> list[DimensionControlRow]:
        """Build DCL rows from raw extracted items when build_dimension_control_list wasn't called."""
        rows: list[DimensionControlRow] = []
        seq = 1
        for pr in result.page_results:
            for dim in pr.dimensions:
                spec = str(dim.specification or dim.value)
                rows.append(DimensionControlRow(
                    dimension_number=seq,
                    specification=spec,
                    criticality=dim.criticality.value if hasattr(dim.criticality, 'value') else "Non-Critical",
                    mode_of_control=dim.mode_of_control.value if hasattr(dim.mode_of_control, 'value') else "Not Defined",
                    mode_of_inspection=dim.mode_of_inspection.value if hasattr(dim.mode_of_inspection, 'value') else "Not Defined",
                    page_number=dim.page_number,
                    category="Dimension",
                    original_id=dim.id,
                ))
                seq += 1
            for tol in pr.tolerances:
                spec = str(tol.specification or tol.value)
                rows.append(DimensionControlRow(
                    dimension_number=seq,
                    specification=spec,
                    criticality=tol.criticality.value if hasattr(tol.criticality, 'value') else "Non-Critical",
                    mode_of_control=tol.mode_of_control.value if hasattr(tol.mode_of_control, 'value') else "Not Defined",
                    mode_of_inspection=tol.mode_of_inspection.value if hasattr(tol.mode_of_inspection, 'value') else "Not Defined",
                    page_number=tol.page_number,
                    category="Tolerance",
                    original_id=tol.id,
                ))
                seq += 1
            for hole in pr.holes:
                spec = str(hole.specification or hole.value)
                rows.append(DimensionControlRow(
                    dimension_number=seq,
                    specification=spec,
                    criticality=hole.criticality.value if hasattr(hole.criticality, 'value') else "Non-Critical",
                    mode_of_control=hole.mode_of_control.value if hasattr(hole.mode_of_control, 'value') else "Not Defined",
                    mode_of_inspection=hole.mode_of_inspection.value if hasattr(hole.mode_of_inspection, 'value') else "Not Defined",
                    page_number=hole.page_number,
                    category="Hole",
                    original_id=hole.id,
                ))
                seq += 1
        return rows

    def _generate_json_report(
        self,
        result: DocumentAnalysisResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        return self.generate_json(result, output_path)
