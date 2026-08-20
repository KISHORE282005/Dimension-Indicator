"""Excel report writer for the fixed-column part report.

The primary sheet always carries exactly these headers, in this order, whatever
the uploaded drawing contains:

    S No | PART NO | DESCRIPTION | DWG NO | WEIGHT (IN KG) |
    THICKNESS | PROCESS | LENGTH (mm) | WIDTH (mm) | HEIGHT (mm)

Three supporting sheets carry everything that does not belong in those ten
columns: per-cell provenance, the raw drawing findings, and the page-by-page
processing log. The report sheet stays clean; the audit trail stays complete.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.models.part_schemas import (
    MM_COLUMNS,
    NOT_AVAILABLE,
    NOT_DETECTED,
    NUMERIC_COLUMNS,
    REPORT_COLUMNS,
    PartReportResult,
    ValueSource,
    ValueStatus,
)

logger = logging.getLogger(__name__)

# --- palette ---------------------------------------------------------------
NAVY = "1F3864"
HEADER_BG = "2F5496"
TITLE_BG = "1F3864"
BAND_BG = "F2F5FA"
CONFLICT_BG = "FFF2CC"
CONFLICT_FG = "7F6000"
MISSING_BG = "FCE4E4"
MISSING_FG = "9C1F1F"
FILLED_BG = "E2EFDA"
FILLED_FG = "1E5B2A"
CONFIRMED_FG = "1E5B2A"
MUTED = "595959"

THIN = Side(style="thin", color="B4C6E7")
MEDIUM = Side(style="medium", color=NAVY)

CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

#: Width in characters for each fixed column.
COLUMN_WIDTHS: dict[str, int] = {
    "S No": 6,
    "PART NO": 14,
    "DESCRIPTION": 40,
    "DWG NO": 18,
    "WEIGHT (IN KG)": 13,
    "THICKNESS": 11,
    "PROCESS": 22,
    "LENGTH (mm)": 12,
    "WIDTH (mm)": 12,
    "HEIGHT (mm)": 12,
}

_NUMBER_ONLY = re.compile(r"^[-+]?\d+(?:[.,]\d+)?$")

#: "6 mm", "2.40kg", '12 "' - a single number followed by a simple unit.
#: Deliberately restricted to known units so that "200 x 100" and "6 to 8"
#: fall through and stay text rather than being silently truncated.
_NUMBER_WITH_UNIT = re.compile(
    r"^(?P<number>[-+]?\d+(?:[.,]\d+)?)\s*"
    r"(?P<unit>mm|cm|m|in|inch|inches|\"|kg|kgs|g|gram|grams|lb|lbs|t|deg|°)\.?$",
    re.IGNORECASE,
)


class ExcelReportWriter:
    """Writes a formatted .xlsx report from a PartReportResult."""

    def __init__(self, output_dir: Optional[Path] = None) -> None:
        self.output_dir = Path(output_dir or settings.OUTPUT_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def build_dataframe(self, result: PartReportResult) -> pd.DataFrame:
        """The report sheet as a DataFrame with the fixed columns in order."""
        records = [row.to_flat_dict() for row in result.rows]
        # Passing an explicit column list guarantees the header contract even
        # when there are no rows at all.
        return pd.DataFrame(records, columns=list(REPORT_COLUMNS))

    def generate(
        self,
        result: PartReportResult,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Write the workbook and return its path."""
        if output_path is None:
            stem = result.document_id[:8] or result.job_id[:8]
            output_path = self.output_dir / f"drawing_report_{stem}.xlsx"
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        df = self.build_dataframe(result)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # startrow leaves space for the title block written afterwards.
            df.to_excel(writer, sheet_name="Report", index=False, startrow=3)
            self._write_traceability(writer, result)
            self._write_findings(writer, result)
            self._write_log(writer, result)

            self._format_report_sheet(writer.book["Report"], result, len(df))

        logger.info(
            "Excel report written to %s (%d part rows)", output_path, len(df)
        )
        return output_path

    # ------------------------------------------------------------------
    # Sheet 1: Report
    # ------------------------------------------------------------------

    def _format_report_sheet(
        self, ws: Worksheet, result: PartReportResult, row_count: int
    ) -> None:
        n_cols = len(REPORT_COLUMNS)
        last_col = get_column_letter(n_cols)
        header_row = 4
        first_data_row = header_row + 1
        last_data_row = header_row + row_count

        # --- Title block ----------------------------------------------
        ws.merge_cells(f"A1:{last_col}1")
        title = ws["A1"]
        title.value = "ENGINEERING DRAWING ANALYSIS REPORT"
        title.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
        title.fill = PatternFill("solid", start_color=TITLE_BG, end_color=TITLE_BG)
        title.alignment = CENTER
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{last_col}2")
        subtitle = ws["A2"]
        subtitle.value = (
            f"{settings.REPORT_COMPANY_NAME}    |    "
            f"Source: {result.filename or 'n/a'}    |    "
            f"Pages analysed: {result.pages_analyzed}/{result.total_pages}    |    "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        subtitle.font = Font(size=9, color=MUTED, italic=True, name="Calibri")
        subtitle.alignment = CENTER
        ws.row_dimensions[2].height = 16

        ws.merge_cells(f"A3:{last_col}3")
        status = ws["A3"]
        status.value = self._status_line(result)
        status.font = Font(size=9, color=MUTED, name="Calibri")
        status.alignment = CENTER
        ws.row_dimensions[3].height = 16

        # --- Header row -----------------------------------------------
        header_fill = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
        for col_index, column in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_index)
            cell.value = column
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            cell.fill = header_fill
            cell.alignment = CENTER
            cell.border = HEADER_BORDER
            ws.column_dimensions[get_column_letter(col_index)].width = (
                COLUMN_WIDTHS.get(column, 14)
            )
        ws.row_dimensions[header_row].height = 28

        # --- Data rows ------------------------------------------------
        for offset, report_row in enumerate(result.rows):
            excel_row = first_data_row + offset
            banded = offset % 2 == 1
            ws.row_dimensions[excel_row].height = 20

            for col_index, column in enumerate(REPORT_COLUMNS, start=1):
                cell = ws.cell(row=excel_row, column=col_index)
                report_cell = report_row.cells.get(column)
                raw_value = report_cell.value if report_cell else NOT_AVAILABLE

                self._write_value(cell, raw_value, column)
                cell.border = CELL_BORDER
                cell.font = Font(size=10, name="Calibri")

                if column in NUMERIC_COLUMNS:
                    cell.alignment = (
                        RIGHT if isinstance(cell.value, (int, float)) else CENTER
                    )
                elif column in {"S No", "PART NO", "DWG NO"}:
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT

                self._apply_status_style(cell, report_cell, banded)

                if report_cell and report_cell.note:
                    cell.comment = self._make_comment(report_cell)

        # --- Freeze, filter, print ------------------------------------
        ws.freeze_panes = ws[f"A{first_data_row}"]
        if row_count:
            ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"

        ws.sheet_properties.tabColor = NAVY
        ws.print_title_rows = f"{header_row}:{header_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # --- Legend ----------------------------------------------------
        legend_row = last_data_row + 2
        ws.merge_cells(f"A{legend_row}:{last_col}{legend_row}")
        legend = ws.cell(row=legend_row, column=1)
        legend.value = (
            "Cell shading -  no fill: value as entered by user  |  "
            "green: filled in from the drawing  |  "
            "amber: your entry differs from the drawing (your value kept)  |  "
            "red: not supplied and not detected.  "
            "Hover any shaded cell for the page reference and confidence."
        )
        legend.font = Font(size=8, italic=True, color=MUTED, name="Calibri")
        legend.alignment = LEFT
        ws.row_dimensions[legend_row].height = 24

    def _write_value(self, cell: Any, raw_value: str, column: str) -> None:
        """Write a value, coercing to a real number where the column expects one."""
        text = (raw_value or "").strip()

        if column in NUMERIC_COLUMNS and text not in {NOT_DETECTED, NOT_AVAILABLE, ""}:
            parsed = self._as_number(text)
            if parsed is not None:
                number, unit = parsed
                cell.value = number
                cell.number_format = self._number_format(column, unit)
                return

        cell.value = text or NOT_AVAILABLE

    @staticmethod
    def _number_format(column: str, unit: str) -> str:
        """Number format that keeps the cell numeric but still shows the unit.

        Writing "200 mm" as a string would make the column unsortable and
        unusable in a formula, while dropping the unit outright would lose
        engineering meaning. A suffixed format keeps both.
        """
        base = "0.00" if column == "WEIGHT (IN KG)" else "0.###"
        # The header already states the unit for these columns, so repeating it
        # inside the cell is noise.
        header_unit = (
            column == "WEIGHT (IN KG)" and unit.lower() in {"kg", "kgf", "kgs"}
        ) or (column in MM_COLUMNS and unit.lower() in {"mm", "millimetre", "millimeter"})
        if not unit or header_unit:
            return base

        # The suffix is embedded in a quoted section of the format string, so a
        # unit containing a double quote (the inch mark) would terminate that
        # section and produce a file Excel offers to "repair". Spell it out.
        suffix = unit.replace('"', "in").replace(";", "")
        if not suffix:
            return base
        return f'{base} "{suffix}"'

    @staticmethod
    def _as_number(text: str) -> Optional[tuple[float, str]]:
        """Split "6", "6 mm" or "2.40kg" into (number, unit).

        Returns None for anything else - "200 x 100" and "6 to 8" stay text,
        because reducing them to a single number would corrupt the value.
        """
        candidate = text.strip()
        if _NUMBER_ONLY.match(candidate.replace(",", ".")):
            try:
                return float(candidate.replace(",", ".")), ""
            except ValueError:
                return None

        match = _NUMBER_WITH_UNIT.match(candidate)
        if not match:
            return None
        try:
            number = float(match.group("number").replace(",", "."))
        except ValueError:
            return None
        return number, match.group("unit").strip()

    @staticmethod
    def _apply_status_style(cell: Any, report_cell: Any, banded: bool) -> None:
        if report_cell is None:
            return

        status = report_cell.status
        if status == ValueStatus.CONFLICT:
            cell.fill = PatternFill("solid", start_color=CONFLICT_BG, end_color=CONFLICT_BG)
            cell.font = Font(size=10, name="Calibri", color=CONFLICT_FG, bold=True)
        elif status == ValueStatus.MISSING:
            cell.fill = PatternFill("solid", start_color=MISSING_BG, end_color=MISSING_BG)
            cell.font = Font(size=10, name="Calibri", color=MISSING_FG, italic=True)
        elif status == ValueStatus.FILLED_FROM_DRAWING and report_cell.source == ValueSource.DRAWING:
            cell.fill = PatternFill("solid", start_color=FILLED_BG, end_color=FILLED_BG)
            cell.font = Font(size=10, name="Calibri", color=FILLED_FG)
        elif status == ValueStatus.CONFIRMED:
            cell.font = Font(size=10, name="Calibri", color=CONFIRMED_FG)
            if banded:
                cell.fill = PatternFill("solid", start_color=BAND_BG, end_color=BAND_BG)
        elif banded:
            cell.fill = PatternFill("solid", start_color=BAND_BG, end_color=BAND_BG)

    @staticmethod
    def _make_comment(report_cell: Any) -> Any:
        from openpyxl.comments import Comment

        lines = [f"Source: {report_cell.source.value}", f"Status: {report_cell.status.value}"]
        if report_cell.user_value:
            lines.append(f"You entered: {report_cell.user_value}")
        if report_cell.drawing_value:
            lines.append(f"Drawing shows: {report_cell.drawing_value}")
        if report_cell.page_references:
            lines.append(
                "Page(s): " + ", ".join(str(p) for p in report_cell.page_references)
            )
        if report_cell.confidence:
            lines.append(f"Confidence: {report_cell.confidence:.0%}")
        if report_cell.note:
            lines.append("")
            lines.append(report_cell.note)

        comment = Comment("\n".join(lines), "Drawing Analysis")
        comment.width = 320
        comment.height = 160
        return comment

    @staticmethod
    def _status_line(result: PartReportResult) -> str:
        bits = [
            f"OCR: {result.ocr_engine}",
            f"Vision model: {result.vlm_model}",
            f"Conflicts: {result.conflict_count()}",
            f"Filled from drawing: {result.filled_count()}",
            f"Not detected: {result.missing_count()}",
        ]
        return "    |    ".join(bits)

    # ------------------------------------------------------------------
    # Sheet 2: Traceability
    # ------------------------------------------------------------------

    def _write_traceability(self, writer: Any, result: PartReportResult) -> None:
        records = []
        for row in result.rows:
            for column in REPORT_COLUMNS:
                cell = row.cells.get(column)
                if cell is None:
                    continue
                records.append(
                    {
                        "Part No": row.part_no,
                        "Column": column,
                        "Reported Value": cell.value,
                        "Source": cell.source.value,
                        "Status": cell.status.value,
                        "User Entered": cell.user_value or "",
                        "Drawing Shows": cell.drawing_value or "",
                        "Page(s)": ", ".join(str(p) for p in cell.page_references),
                        "Confidence": round(cell.confidence, 3),
                        "Note": cell.note or "",
                    }
                )

        df = pd.DataFrame(
            records,
            columns=[
                "Part No", "Column", "Reported Value", "Source", "Status",
                "User Entered", "Drawing Shows", "Page(s)", "Confidence", "Note",
            ],
        )
        df.to_excel(writer, sheet_name="Traceability", index=False)
        self._format_simple_sheet(
            writer.book["Traceability"],
            widths=[14, 16, 22, 12, 16, 16, 18, 10, 11, 55],
            percent_columns={9},
        )

    # ------------------------------------------------------------------
    # Sheet 3: Drawing Information
    # ------------------------------------------------------------------

    def _write_findings(self, writer: Any, result: PartReportResult) -> None:
        records = []
        for row in result.rows:
            for finding in row.findings:
                records.append(self._finding_record(finding, row.part_no))
        for finding in result.unmatched_findings:
            records.append(self._finding_record(finding, "(unattributed)"))

        records.sort(key=lambda r: (r["Page"], r["Category"]))

        df = pd.DataFrame(
            records,
            columns=["Page", "Part No", "Category", "Value", "Detail", "Confidence", "Source"],
        )
        df.to_excel(writer, sheet_name="Drawing Information", index=False)
        self._format_simple_sheet(
            writer.book["Drawing Information"],
            widths=[7, 18, 16, 40, 50, 11, 10],
            percent_columns={5},
        )

    @staticmethod
    def _finding_record(finding: Any, part_no: str) -> dict:
        return {
            "Page": finding.page_number,
            "Part No": finding.part_no or part_no,
            "Category": finding.category.replace("_", " ").title(),
            "Value": finding.value,
            "Detail": finding.detail or "",
            "Confidence": round(finding.confidence, 3),
            "Source": finding.source,
        }

    # ------------------------------------------------------------------
    # Sheet 4: Analysis Log
    # ------------------------------------------------------------------

    def _write_log(self, writer: Any, result: PartReportResult) -> None:
        records = [
            {
                "Page": p.page_number,
                "OCR Engine": p.ocr_engine,
                "Text Regions": p.ocr_regions,
                "Characters Read": p.ocr_text_length,
                "Vision Model Used": "yes" if p.vlm_used else "no",
                "Items Found": p.findings_count,
                "Part Numbers Seen": ", ".join(p.part_numbers_seen),
                "Seconds": round(p.processing_time_seconds, 2),
                "Error": p.vlm_error or "",
            }
            for p in result.page_analyses
        ]
        df = pd.DataFrame(
            records,
            columns=[
                "Page", "OCR Engine", "Text Regions", "Characters Read",
                "Vision Model Used", "Items Found", "Part Numbers Seen",
                "Seconds", "Error",
            ],
        )
        df.to_excel(writer, sheet_name="Analysis Log", index=False)
        ws = writer.book["Analysis Log"]
        self._format_simple_sheet(ws, widths=[7, 16, 13, 16, 17, 12, 26, 10, 45])

        # Append run-level messages beneath the per-page table.
        next_row = len(records) + 3
        for label, messages in (("Warnings", result.warnings), ("Errors", result.errors)):
            if not messages:
                continue
            header = ws.cell(row=next_row, column=1)
            header.value = label
            header.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
            next_row += 1
            for message in messages:
                ws.merge_cells(
                    start_row=next_row, start_column=1, end_row=next_row, end_column=9
                )
                cell = ws.cell(row=next_row, column=1)
                cell.value = message
                cell.alignment = TOP_LEFT
                cell.font = Font(
                    size=9, name="Calibri",
                    color=MISSING_FG if label == "Errors" else CONFLICT_FG,
                )
                ws.row_dimensions[next_row].height = 26
                next_row += 1
            next_row += 1

    # ------------------------------------------------------------------
    # Shared formatting
    # ------------------------------------------------------------------

    @staticmethod
    def _format_simple_sheet(
        ws: Worksheet,
        widths: list[int],
        percent_columns: Optional[set[int]] = None,
    ) -> None:
        """Apply the house style to a plain header-plus-rows sheet."""
        percent_columns = percent_columns or set()
        header_fill = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
        max_row = ws.max_row
        max_col = ws.max_column

        for col_index in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_index)
            cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill = header_fill
            cell.alignment = CENTER
            cell.border = HEADER_BORDER
            width = widths[col_index - 1] if col_index - 1 < len(widths) else 16
            ws.column_dimensions[get_column_letter(col_index)].width = width
        ws.row_dimensions[1].height = 24

        for row_index in range(2, max_row + 1):
            banded = row_index % 2 == 1
            for col_index in range(1, max_col + 1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.font = Font(size=9, name="Calibri")
                cell.border = CELL_BORDER
                cell.alignment = TOP_LEFT
                if banded:
                    cell.fill = PatternFill("solid", start_color=BAND_BG, end_color=BAND_BG)
                if (col_index - 1) in percent_columns and isinstance(
                    cell.value, (int, float)
                ):
                    cell.number_format = "0%"
                    cell.alignment = CENTER

        ws.freeze_panes = "A2"
        if max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
